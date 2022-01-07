from openpyxl import load_workbook 
from openpyxl.workbook import workbook
import os


root_directory = os.path.dirname(os.path.abspath(__file__))

data_name = input("\nEnter name of export shipping data file: ")
data_name = data_name.strip() 

report_name = input("\nEnter name of container report: ")
report_name = report_name.strip()

data_file_path = "{}/{}".format(root_directory, data_name)
report_file_path = "{}/{}".format(root_directory, report_name)

data_file_found = os.path.isfile(data_file_path)
report_file_found = os.path.isfile(report_file_path)


if data_file_found and report_file_found:
    data_workbook = load_workbook(data_file_path)
    data_ws = data_workbook.active

    report_workbook = load_workbook(report_file_path)

    def check_container_sheet():
        for i in report_workbook:
            if i.title == "CONTAINERS":
                return True

        return False

    if check_container_sheet():
        report_ws = report_workbook['CONTAINERS']

        def find_last_index():
            for i in range(report_ws.max_row,1,-1):
                if report_ws['A{}'.format(i)].value == "TOTAL UNIT":
                    return i
            return report_ws.max_row

        if data_ws['A1'].value == "Booking Office Code" and report_ws['A1'].value == "CONTAINER PICK UP AREA":
            shipping_lane = ""
            for i in range(2,find_last_index()):
                report_ship_names = report_ws['A{}'.format(i)].value
                
                if isinstance(report_ship_names, str):
                    ignore_list = ['CPNW', 'CENX', 'MPNW', 'OPNW','EPNW', 'BLANK']
                    report_ship_name_list = report_ship_names.split(" ")
                    if not report_ship_name_list[0].strip() in ignore_list:
                        char_code = report_ws['B{}'.format(i)].value
                        number_code = report_ship_name_list[len(report_ship_name_list)-1]
                        shipping_code = "{}-{}-{}".format(shipping_lane, char_code, number_code) 

                        data = {
                            "VAN/PRR": {
                                "20GP": 0,
                                "40GP": 0,
                                "40HQ": 0,
                                "40RQ": 0,
                                "45": 0,
                            },

                            "Prince George": {
                                "40GP": 0,
                                "40HQ": 0,
                                "40RQ": 0,
                            },

                            "Calgary": {
                                "40GP": 0,
                                "40HQ": 0,
                                "40RQ": 0,
                            },

                            "Edmonton": {
                                "20GP": 0,
                                "40GP": 0,
                                "40HQ": 0,
                                "40RQ": 0,
                            },
                        }

                        for j in range(2, data_ws.max_row):
                            data_code = data_ws['F{}'.format(j)].value.split(" ")[:-1][0]
                            if shipping_code == data_code:
                                
                                origin = data_ws['D{}'.format(j)].value.split(",")[0].strip()
                                if origin != "Montreal" and origin != "Toronto" and origin != "Halifax":
                                    GP_20 = data_ws['H{}'.format(j)].value
                                    GP_40 = data_ws['I{}'.format(j)].value
                                    HQ_40 = data_ws['J{}'.format(j)].value
                                    A_45 = data_ws['L{}'.format(j)].value
                                    RQ_40 = data_ws['K{}'.format(j)].value
                                    if origin == "Vancouver" or origin == "Prince Rupert":
                                       data["VAN/PRR"]["20GP"] = data["VAN/PRR"]["20GP"] + GP_20
                                       data["VAN/PRR"]["40GP"] = data["VAN/PRR"]["40GP"] + GP_40
                                       data["VAN/PRR"]["40HQ"] = data["VAN/PRR"]["40HQ"] + HQ_40
                                       data["VAN/PRR"]["45"] = data["VAN/PRR"]["45"] + A_45
                                       data["VAN/PRR"]["40RQ"] = data["VAN/PRR"]["40RQ"] + RQ_40
                                    else:
                                        
                                        for z in data[origin]:
                                            match = {
                                                "20GP": GP_20,
                                                "40GP": GP_40,
                                                "40HQ": HQ_40,
                                                "40RQ": RQ_40,
                                                "45": A_45,
                                            }

                                            data[origin][z] = data[origin][z] + match[z]

                                      
                        for k in data:
                            if k == "VAN/PRR":
                                counter = 0
                                options = ['20GP', '40GP', '40HQ', '40RQ', '45']
                                for j in ['C','D','E','F','H']:
                                    if data[k][options[counter]] != 0:
                                        report_ws['{}{}'.format(j,i)] = data[k][options[counter]]
                                    else:
                                        report_ws['{}{}'.format(j,i)] = ""

                                    counter +=1
                            elif k == "Prince George":
                                counter = 0
                                options = ['40GP', '40HQ', '40RQ']
                                for j in ['I','J','K']:
                                    if data[k][options[counter]] != 0:
                                        report_ws['{}{}'.format(j,i)] = data[k][options[counter]]
                                    else:
                                        report_ws['{}{}'.format(j,i)] = ""

                                    counter +=1
                            elif k == "Calgary":
                                if i == 22:
                                    print(data) 
                                counter = 0
                                options = ['40GP', '40HQ', '40RQ']
                                for j in ['L','M','N']:
                                    if data[k][options[counter]] != 0:
                                        report_ws['{}{}'.format(j,i)] = data[k][options[counter]]
                                    else:
                                        
                                        report_ws['{}{}'.format(j,i)] = 0
                                        if i == 22:
                                            print(str(report_ws['{}{}'.format(j, i)].value) + " " + '{} {}'.format(j,i))
                                        
                                    counter +=1

                            elif k == "Edmonton":
                                counter = 0
                                options = ['20GP', '40GP', '40HQ', '40RQ']
                                for j in ['O','P','Q','R']:
                                    if data[k][options[counter]] != 0:
                                        report_ws['{}{}'.format(j,i)] = data[k][options[counter]]
                                    else:
                                        report_ws['{}{}'.format(j,i)] = ""

                                    counter +=1
                                



                        
                    elif report_ship_name_list[0].strip() != "BLANK":
                        shipping_lane = report_ship_name_list[0].strip()  
            report_workbook.save('Container Report.xlsx')  
            print("\n Process Complete \n")

            #new_file_name = input("\n What would you like to name your updated file?: ")
            #if new_file_name.split(".")[1] == "xlsx":
                #os.rename(report_file_path, "{}/{}".format(root_directory, new_file_name))
                #os.remove(data_file_path)
                #print("\n Success! Your new file is ready, please remove from folder directory.")

            #else:
                #print("\n make sure to rename your file with the .xlsx at the end")
                            
        else:
            print("Error")
            








